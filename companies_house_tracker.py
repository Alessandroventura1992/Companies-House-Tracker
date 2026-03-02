"""
UK Company Incorporation Tracker
Polls the Companies House API every 30 minutes for newly incorporated companies
by SIC code category, exports results to Excel with director cross-referencing.
"""

import base64
import json
import logging
import os
import sys
import time
from datetime import date, datetime, timedelta
from logging.handlers import RotatingFileHandler
from pathlib import Path

import requests
import schedule
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE_DIR = Path(__file__).parent
CONFIG_PATH = BASE_DIR / "config.json"
STATE_PATH = BASE_DIR / "tracker_state.json"
DATA_STORE_PATH = BASE_DIR / "data_store.json"
LOG_PATH = BASE_DIR / "tracker.log"

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
def setup_logging():
    logger = logging.getLogger("tracker")
    logger.setLevel(logging.DEBUG)
    fmt = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", "%Y-%m-%d %H:%M:%S")

    fh = RotatingFileHandler(LOG_PATH, maxBytes=5 * 1024 * 1024, backupCount=3, encoding="utf-8")
    fh.setFormatter(fmt)
    fh.setLevel(logging.DEBUG)

    ch = logging.StreamHandler(sys.stdout)
    ch.setFormatter(fmt)
    ch.setLevel(logging.INFO)

    logger.addHandler(fh)
    logger.addHandler(ch)
    return logger

log = setup_logging()

# ---------------------------------------------------------------------------
# Excel colour palette
# ---------------------------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")   # dark blue
ALT_ROW_FILL = PatternFill("solid", fgColor="D6E4F0")  # light blue
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)

# Sheet tab colours per category index
TAB_COLOURS = [
    "4472C4", "ED7D31", "A9D18E", "FF0000",
    "FFC000", "70AD47", "5A96C8", "C55A11",
    "00B0F0", "7030A0", "92D050",
]

# ---------------------------------------------------------------------------
# Main Tracker class
# ---------------------------------------------------------------------------
class CompaniesHouseTracker:

    BASE_URL = "https://api.company-information.service.gov.uk"

    def __init__(self, config_path: Path = CONFIG_PATH):
        self.config = self._load_json(config_path, required=True)
        self.api_key = os.environ.get("COMPANIES_HOUSE_API_KEY") or self.config.get("api_key", "")
        self.poll_interval = self.config.get("poll_interval_minutes", 30)
        self.initial_lookback = self.config.get("initial_lookback_days", 7)
        self.output_path = BASE_DIR / self.config.get("output_filename", "companies_tracker.xlsx")
        self.sic_categories: dict[str, list[str]] = self.config["sic_categories"]

        auth_raw = base64.b64encode(f"{self.api_key}:".encode()).decode()
        self.headers = {"Authorization": f"Basic {auth_raw}"}

        self.state = self._load_json(STATE_PATH, default={"last_run": None, "seen": []})
        self.data_store: dict = self._load_json(DATA_STORE_PATH, default={})

        self._last_request_time = 0.0

    # ------------------------------------------------------------------
    # JSON helpers
    # ------------------------------------------------------------------
    def _load_json(self, path: Path, required=False, default=None):
        if path.exists():
            try:
                return json.loads(path.read_text(encoding="utf-8"))
            except json.JSONDecodeError as e:
                log.error("Corrupt JSON at %s: %s", path, e)
                if required:
                    sys.exit(1)
        if required:
            log.error("Required config file not found: %s", path)
            sys.exit(1)
        return default if default is not None else {}

    def _save_json(self, path: Path, data):
        path.write_text(json.dumps(data, indent=2, default=str), encoding="utf-8")

    # ------------------------------------------------------------------
    # API
    # ------------------------------------------------------------------
    def _rate_limit(self):
        """Ensure at least 0.6s between requests (~1.67 req/sec)."""
        elapsed = time.monotonic() - self._last_request_time
        if elapsed < 0.6:
            time.sleep(0.6 - elapsed)
        self._last_request_time = time.monotonic()

    def _make_request(self, endpoint: str, params: dict = None, retries: int = 3) -> dict | None:
        url = self.BASE_URL + endpoint
        for attempt in range(retries):
            self._rate_limit()
            try:
                resp = requests.get(url, headers=self.headers, params=params, timeout=15)
                if resp.status_code == 200:
                    return resp.json()
                if resp.status_code == 429:
                    wait = 65
                    log.warning("Rate limited (429). Waiting %ss before retry...", wait)
                    time.sleep(wait)
                    continue
                if resp.status_code == 404:
                    log.debug("404 for %s", url)
                    return None
                if resp.status_code == 401:
                    log.error("Authentication failed (401). Check your API key in config.json.")
                    return None
                log.warning("HTTP %s for %s (attempt %s/%s)", resp.status_code, url, attempt + 1, retries)
            except requests.RequestException as e:
                log.warning("Request error (attempt %s/%s): %s", attempt + 1, retries, e)
                time.sleep(5)
        log.error("All retries exhausted for %s", endpoint)
        return None

    def _search_by_sic(self, sic_code: str, from_date: str, to_date: str) -> list[dict]:
        """Return all companies for a given SIC code incorporated in [from_date, to_date]."""
        results = []
        start_index = 0
        while True:
            params = {
                "sic_codes": sic_code,
                "incorporated_from": from_date,
                "incorporated_to": to_date,
                "size": 100,
                "start_index": start_index,
            }
            data = self._make_request("/advanced-search/companies", params)
            if not data:
                break
            items = data.get("items", [])
            if not items:
                break
            results.extend(items)
            total_hits = data.get("hits", 0)
            start_index += len(items)
            if start_index >= total_hits:
                break
        return results

    def _get_officers(self, company_number: str) -> list[dict]:
        """Return active directors for a company."""
        data = self._make_request(f"/company/{company_number}/officers", {"items_per_page": 100})
        if not data:
            return []
        officers = []
        for item in data.get("items", []):
            if item.get("resigned_on"):
                continue
            role = item.get("officer_role", "")
            if "director" not in role.lower() and "llp-member" not in role.lower():
                continue
            officers.append({
                "name": item.get("name", "Unknown"),
                "role": role,
                "appointed_on": item.get("appointed_on", ""),
                "nationality": item.get("nationality", ""),
                "country_of_residence": item.get("country_of_residence", ""),
                "date_of_birth": item.get("date_of_birth", {}),
            })
        return officers

    def _get_pscs(self, company_number: str) -> list[dict]:
        """Return persons with significant control (beneficial owners)."""
        data = self._make_request(
            f"/company/{company_number}/persons-with-significant-control",
            {"items_per_page": 100},
        )
        if not data:
            return []
        pscs = []
        for item in data.get("items", []):
            if item.get("ceased_on"):
                continue
            kind = item.get("kind", "")
            name = item.get("name", "Unknown")
            controls = "; ".join(item.get("natures_of_control", []))
            dob = item.get("date_of_birth", {})
            pscs.append({
                "name": name,
                "kind": kind,
                "natures_of_control": controls,
                "nationality": item.get("nationality", ""),
                "country_of_residence": item.get("country_of_residence", ""),
                "date_of_birth": dob,
            })
        return pscs

    # ------------------------------------------------------------------
    # Core fetch logic
    # ------------------------------------------------------------------
    def fetch_new_companies(self) -> int:
        """
        Search for newly incorporated companies across all SIC categories.
        Only processes companies not already in data_store (seen set).
        Returns count of new companies added.
        """
        seen: set[str] = set(self.state.get("seen", []))

        today_str = date.today().isoformat()
        last_run = self.state.get("last_run")
        if last_run:
            from_date = today_str  # same-day only on subsequent runs
        else:
            lookback = date.today() - timedelta(days=self.initial_lookback)
            from_date = lookback.isoformat()
        to_date = today_str

        log.info("Searching incorporations from %s to %s", from_date, to_date)

        # Collect all unique company numbers across ALL SIC codes first,
        # then fetch details only for unseen ones.
        all_raw: dict[str, dict] = {}  # company_number → raw API item
        for category, sic_codes in self.sic_categories.items():
            unique_sics = list(dict.fromkeys(sic_codes))  # preserve order, remove dupes
            for sic in unique_sics:
                log.info("  Searching SIC %s (%s)...", sic, category)
                items = self._search_by_sic(sic, from_date, to_date)
                for item in items:
                    num = item.get("company_number", "")
                    if num and num not in all_raw:
                        all_raw[num] = item
                log.debug("    SIC %s → %s result(s)", sic, len(items))

        new_numbers = [n for n in all_raw if n not in seen]
        log.info("Found %s unique companies total, %s are new.", len(all_raw), len(new_numbers))

        added = 0
        for i, company_number in enumerate(new_numbers, 1):
            raw = all_raw[company_number]
            log.info("  [%s/%s] Fetching details: %s (%s)", i, len(new_numbers),
                     raw.get("company_name", "?"), company_number)
            officers = self._get_officers(company_number)
            pscs = self._get_pscs(company_number)

            address_obj = raw.get("registered_office_address", {})
            address = ", ".join(filter(None, [
                address_obj.get("address_line_1", ""),
                address_obj.get("address_line_2", ""),
                address_obj.get("locality", ""),
                address_obj.get("postal_code", ""),
            ]))

            self.data_store[company_number] = {
                "company_number": company_number,
                "company_name": raw.get("company_name", "Unknown"),
                "date_of_creation": raw.get("date_of_creation", ""),
                "company_status": raw.get("company_status", ""),
                "sic_codes": raw.get("sic_codes", []),
                "registered_address": address,
                "officers": officers,
                "pscs": pscs,
                "fetched_at": datetime.now().isoformat(),
            }
            seen.add(company_number)
            added += 1

        self.state["last_run"] = datetime.now().isoformat()
        self.state["seen"] = list(seen)
        self._save_json(STATE_PATH, self.state)
        self._save_json(DATA_STORE_PATH, self.data_store)
        log.info("Added %s new companies. Data store now has %s companies.", added, len(self.data_store))
        return added

    # ------------------------------------------------------------------
    # Director linking
    # ------------------------------------------------------------------
    def _build_director_links(self) -> dict[str, list[str]]:
        """
        Returns {director_key: [company_number, ...]} for all companies.
        director_key = lowercased name + "_" + dob month + "_" + dob year
        """
        director_map: dict[str, list[str]] = {}
        for company_number, company in self.data_store.items():
            for officer in company.get("officers", []):
                dob = officer.get("date_of_birth", {})
                key = (
                    officer["name"].lower().strip()
                    + "_"
                    + str(dob.get("month", ""))
                    + "_"
                    + str(dob.get("year", ""))
                )
                director_map.setdefault(key, [])
                if company_number not in director_map[key]:
                    director_map[key].append(company_number)
        return director_map

    # ------------------------------------------------------------------
    # Excel export
    # ------------------------------------------------------------------
    def _format_directors(self, officers: list[dict]) -> str:
        if not officers:
            return "None listed"
        parts = []
        for o in officers:
            dob = o.get("date_of_birth", {})
            dob_str = f"{dob.get('month', '?')}/{dob.get('year', '?')}" if dob else ""
            nat = o.get("nationality", "")
            detail = " | ".join(filter(None, [dob_str, nat]))
            parts.append(f"{o['name']} [{detail}]" if detail else o["name"])
        return "\n".join(parts)

    def _format_pscs(self, pscs: list[dict]) -> str:
        if not pscs:
            return "Not registered"
        parts = []
        for p in pscs:
            kind = p.get("kind", "")
            label = "Corporate" if "corporate" in kind else "Individual"
            controls = p.get("natures_of_control", "")
            parts.append(f"{p['name']} ({label})" + (f" — {controls}" if controls else ""))
        return "\n".join(parts)

    def _format_linked(self, company_number: str, director_map: dict[str, list[str]]) -> str:
        company = self.data_store.get(company_number, {})
        linked: set[str] = set()
        for officer in company.get("officers", []):
            dob = officer.get("date_of_birth", {})
            key = (
                officer["name"].lower().strip()
                + "_"
                + str(dob.get("month", ""))
                + "_"
                + str(dob.get("year", ""))
            )
            for num in director_map.get(key, []):
                if num != company_number:
                    name = self.data_store.get(num, {}).get("company_name", num)
                    linked.add(f"{name} ({num})")
        return "\n".join(sorted(linked)) if linked else ""

    def _write_sheet(self, ws, companies: list[dict], director_map: dict[str, list[str]]):
        headers = [
            "Company Name", "Company No.", "Incorporated",
            "Registered Address", "SIC Code(s)",
            "Directors", "Beneficial Owners", "Linked Companies (same director)",
        ]
        ws.append(headers)
        # Style header row
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
        ws.row_dimensions[1].height = 28

        for row_num, company in enumerate(companies, 2):
            sic_str = ", ".join(company.get("sic_codes", []))
            directors_str = self._format_directors(company.get("officers", []))
            pscs_str = self._format_pscs(company.get("pscs", []))
            linked_str = self._format_linked(company["company_number"], director_map)

            row_data = [
                company.get("company_name", ""),
                company.get("company_number", ""),
                company.get("date_of_creation", ""),
                company.get("registered_address", ""),
                sic_str,
                directors_str,
                pscs_str,
                linked_str,
            ]
            ws.append(row_data)

            fill = ALT_ROW_FILL if row_num % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.fill = fill
                cell.font = BODY_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="top", wrap_text=True)

            # Auto row height: estimate 15pt per line
            max_lines = max(
                directors_str.count("\n") + 1,
                pscs_str.count("\n") + 1,
                linked_str.count("\n") + 1,
                1,
            )
            ws.row_dimensions[row_num].height = max(15, min(max_lines * 15, 120))

        self._auto_column_widths(ws)
        ws.freeze_panes = "A2"

    def _auto_column_widths(self, ws):
        col_max_widths = {
            1: 35,   # Company Name
            2: 14,   # Company No.
            3: 14,   # Incorporated
            4: 35,   # Address
            5: 20,   # SIC
            6: 45,   # Directors
            7: 45,   # Beneficial Owners
            8: 45,   # Linked Companies
        }
        for col_idx, max_w in col_max_widths.items():
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = max_w

    def _write_cross_reference_sheet(self, ws, director_map: dict[str, list[str]]):
        headers = ["Director Name", "DOB (M/Y)", "No. of Companies", "Company Names", "Company Numbers"]
        ws.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
        ws.row_dimensions[1].height = 28

        # Only show directors linked to 2+ companies
        multi_linked = {k: v for k, v in director_map.items() if len(v) >= 2}
        sorted_entries = sorted(multi_linked.items(), key=lambda x: -len(x[1]))

        for row_num, (key, company_numbers) in enumerate(sorted_entries, 2):
            parts = key.rsplit("_", 2)
            name = parts[0].title() if parts else key
            month = parts[1] if len(parts) > 1 else ""
            year = parts[2] if len(parts) > 2 else ""
            dob_str = f"{month}/{year}" if month or year else ""

            company_names = "\n".join(
                self.data_store.get(n, {}).get("company_name", n) for n in company_numbers
            )
            company_nums = "\n".join(company_numbers)

            ws.append([name, dob_str, len(company_numbers), company_names, company_nums])

            fill = ALT_ROW_FILL if row_num % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_num, column=col_idx)
                cell.fill = fill
                cell.font = BODY_FONT
                cell.border = THIN_BORDER
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            lines = len(company_numbers)
            ws.row_dimensions[row_num].height = max(15, min(lines * 15, 120))

        for col_letter, width in zip("ABCDE", [35, 12, 16, 45, 16]):
            ws.column_dimensions[col_letter].width = width
        ws.freeze_panes = "A2"

    def export_to_excel(self):
        if not self.data_store:
            log.info("No data to export yet.")
            return

        director_map = self._build_director_links()
        wb = Workbook()
        wb.remove(wb.active)  # remove default sheet

        # --- Per-category sheets ---
        category_names = list(self.sic_categories.keys())
        for cat_idx, (category, sic_codes) in enumerate(self.sic_categories.items()):
            sic_set = set(sic_codes)
            companies = [
                c for c in self.data_store.values()
                if sic_set.intersection(set(c.get("sic_codes", [])))
            ]
            companies.sort(key=lambda c: c.get("date_of_creation", ""), reverse=True)

            sheet_name = category[:31]  # Excel 31-char limit
            ws = wb.create_sheet(title=sheet_name)
            ws.sheet_properties.tabColor = TAB_COLOURS[cat_idx % len(TAB_COLOURS)]

            self._write_sheet(ws, companies, director_map)
            log.debug("Sheet '%s': %s companies", sheet_name, len(companies))

        # --- All Companies sheet ---
        all_companies = sorted(self.data_store.values(),
                               key=lambda c: c.get("date_of_creation", ""), reverse=True)
        ws_all = wb.create_sheet(title="All Companies")
        ws_all.sheet_properties.tabColor = "404040"
        self._write_sheet(ws_all, all_companies, director_map)

        # --- Director Cross-Reference sheet ---
        ws_xref = wb.create_sheet(title="Director Cross-Reference")
        ws_xref.sheet_properties.tabColor = "FF0000"
        self._write_cross_reference_sheet(ws_xref, director_map)

        wb.save(self.output_path)
        log.info("Excel saved → %s  (%s companies across %s categories)",
                 self.output_path, len(self.data_store), len(self.sic_categories))

    # ------------------------------------------------------------------
    # Scheduler entry point
    # ------------------------------------------------------------------
    def _run_cycle(self):
        log.info("=" * 60)
        log.info("Starting fetch cycle at %s", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
        try:
            added = self.fetch_new_companies()
            if added > 0 or not self.output_path.exists():
                self.export_to_excel()
            else:
                log.info("No new companies — Excel not regenerated.")
        except Exception as e:
            log.exception("Unexpected error during fetch cycle: %s", e)
        log.info("Cycle complete. Next run in %s minutes.", self.poll_interval)

    def run(self):
        log.info("UK Company Incorporation Tracker started.")
        log.info("Polling every %s minutes. Press Ctrl+C to stop.", self.poll_interval)
        log.info("Output: %s", self.output_path)

        # Run immediately on startup
        self._run_cycle()

        schedule.every(self.poll_interval).minutes.do(self._run_cycle)
        while True:
            schedule.run_pending()
            time.sleep(1)

    def run_once(self):
        """Single fetch-and-export cycle. Used by GitHub Actions."""
        log.info("UK Company Incorporation Tracker — single run mode.")
        log.info("Output: %s", self.output_path)
        self._run_cycle()


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------
if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="UK Company Incorporation Tracker")
    parser.add_argument(
        "--once",
        action="store_true",
        help="Run one cycle and exit (for CI/GitHub Actions)",
    )
    args = parser.parse_args()

    tracker = CompaniesHouseTracker(CONFIG_PATH)
    try:
        if args.once:
            tracker.run_once()
        else:
            tracker.run()
    except KeyboardInterrupt:
        log.info("Tracker stopped by user.")
